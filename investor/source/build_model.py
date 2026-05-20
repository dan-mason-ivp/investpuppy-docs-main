"""
InvestPuppy — Seed Round Financial Model
ip-investor-model.xlsx
4 sheets: Cover · Assumptions · P&L · Cashflow & Runway
Run from repo root: python3 investor/source/build_model.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ── Paths ──────────────────────────────────────────────────────────────────────
HERE    = Path(__file__).resolve().parent
OUT_DIR = HERE.parent / "output" / "xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT     = OUT_DIR / "ip-investor-model.xlsx"

# ── Colour palette (openpyxl uses ARGB hex) ────────────────────────────────────
# Brand
C_GREEN      = "FF85D155"
C_DARK       = "FF0F0F14"
C_WHITE      = "FFFFFFFF"
C_LIGHT_GREY = "FFF6F6F6"
C_MID_GREY   = "FFE2E2E2"
C_PALE_GREEN = "FFEEF7E6"

# Industry-standard financial model colours
C_INPUT_BLUE = "FF0000FF"   # Blue text: hardcoded inputs
C_FORMULA    = "FF000000"   # Black text: formulas
C_LINK_GREEN = "FF008000"   # Green text: cross-sheet links
C_ASSUMPTION_BG = "FFFFF0CC" # Light yellow: key assumptions

# ── Style helpers ──────────────────────────────────────────────────────────────
def font(bold=False, italic=False, size=10, color="FF000000", name="Arial"):
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(top=False, bottom=False, left=False, right=False):
    t = Side(style="thin") if True else None
    n = None
    return Border(
        top=Side(style="thin") if top else n,
        bottom=Side(style="thin") if bottom else n,
        left=Side(style="thin") if left else n,
        right=Side(style="thin") if right else n,
    )

def thick_bottom():
    return Border(bottom=Side(style="medium"))

def set_cell(ws, row, col, value, bold=False, italic=False, size=10,
             color="FF000000", bg=None, h_align="left", v_align="center",
             wrap=False, num_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, italic=italic, size=size, color=color)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h_align, v_align, wrap)
    if num_format:
        c.number_format = num_format
    if border:
        c.border = border
    return c

def header_row(ws, row, cols_values, bg=C_DARK, fg=C_WHITE, size=9, bold=True):
    for col, val in cols_values:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=bold, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = align("center", "center")

def section_header(ws, row, col, text, col_span=None):
    c = ws.cell(row=row, column=col, value=text.upper())
    c.font = font(bold=True, size=8, color=C_GREEN)
    c.border = Border(bottom=Side(style="medium", color="FF85D155"))
    if col_span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col_span)

# Number formats
NF_SGD    = '"S$"#,##0;("S$"#,##0);"-"'
NF_SGD_K  = '"S$"#,##0,"K";("S$"#,##0,"K");"-"'
NF_PCT    = '0.0%;(0.0%);"-"'
NF_NUM    = '#,##0;(#,##0);"-"'
NF_MONTH  = '#,##0;(#,##0);"-"'

# ── Workbook setup ─────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ══════════════════════════════════════════════════════════════════════════════
ws_cov = wb.create_sheet("Cover")
ws_cov.sheet_view.showGridLines = False
ws_cov.column_dimensions["A"].width = 2
ws_cov.column_dimensions["B"].width = 40
ws_cov.column_dimensions["C"].width = 30
ws_cov.column_dimensions["D"].width = 20

# Title block
ws_cov.row_dimensions[2].height = 8
ws_cov.row_dimensions[3].height = 36
ws_cov.row_dimensions[4].height = 22
ws_cov.row_dimensions[5].height = 16
ws_cov.row_dimensions[6].height = 16

set_cell(ws_cov, 3, 2, "INVESTPUPPY — SEED ROUND FINANCIAL MODEL",
         bold=True, size=18, color=C_WHITE, bg=C_DARK, h_align="left", v_align="center")
ws_cov.merge_cells("B3:D3")

set_cell(ws_cov, 4, 2, "Vektor Platform  ·  Illustrative projections  ·  May 2026",
         italic=True, size=10, color="FF999999", bg=C_DARK, h_align="left")
ws_cov.merge_cells("B4:D4")

set_cell(ws_cov, 5, 2, "S$1,000,000 – S$2,000,000 Seed Round",
         bold=True, size=11, color=C_GREEN, bg=C_DARK, h_align="left")
ws_cov.merge_cells("B5:D5")

ws_cov.row_dimensions[6].height = 10
ws_cov.cell(6, 2).fill = fill(C_DARK)
ws_cov.cell(6, 3).fill = fill(C_DARK)
ws_cov.cell(6, 4).fill = fill(C_DARK)

# Summary table
r = 8
set_cell(ws_cov, r, 2, "MODEL OVERVIEW", bold=True, size=8, color=C_GREEN)
r += 1

summary = [
    ("Raise target",            "S$1M–S$2M seed round"),
    ("Revenue model",           "AUM-tiered annual subscription (set)"),
    ("FM pricing — years 1–3 (locked)", "S$18,000/yr per client"),
    ("FM pricing — post year 3 (permanent)", "1 tier below AUM tier; floor Entry S$24,000/yr"),
    ("Revenue target",          "First revenue within 9–13 months of close"),
    ("Operational runway",      "18–24 months at base burn rate"),
    ("Engineering spend (~45%)","Simulation → live-trading transition"),
    ("Sales/BD spend (~30%)",   "Founding Mandate Programme execution"),
    ("Ops runway (~25%)",       "Regulatory, legal, infrastructure"),
]
for label, val in summary:
    set_cell(ws_cov, r, 2, label, bold=True, size=9, color="FF1A1A1A",
             bg=C_LIGHT_GREY if r % 2 == 0 else C_WHITE)
    set_cell(ws_cov, r, 3, val, size=9, color="FF5A5A5A",
             bg=C_LIGHT_GREY if r % 2 == 0 else C_WHITE)
    r += 1

r += 1
set_cell(ws_cov, r, 2, "COLOUR CODING", bold=True, size=8, color=C_GREEN)
r += 1
coding = [
    ("Blue text",   "Hardcoded inputs — change these for scenario analysis"),
    ("Black text",  "Formulas — do not edit"),
    ("Green text",  "Cross-sheet links — do not edit"),
    ("Yellow bg",   "Key assumptions requiring attention or update"),
]
for colour, desc in coding:
    set_cell(ws_cov, r, 2, colour, bold=True, size=9, color="FF1A1A1A",
             bg=C_LIGHT_GREY if r % 2 == 0 else C_WHITE)
    set_cell(ws_cov, r, 3, desc, size=9, color="FF5A5A5A",
             bg=C_LIGHT_GREY if r % 2 == 0 else C_WHITE)
    r += 1

r += 1
set_cell(ws_cov, r, 2,
    "IMPORTANT: All projections are illustrative only and based on stated assumptions. "
    "Revenue model not finalised. Not a financial forecast. Do not distribute without "
    "the accompanying Investment Memorandum.",
    italic=True, size=8, color="FF999999", wrap=True)
ws_cov.merge_cells(f"B{r}:D{r}")
ws_cov.row_dimensions[r].height = 40

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws_ass = wb.create_sheet("Assumptions")
ws_ass.sheet_view.showGridLines = False
ws_ass.column_dimensions["A"].width = 2
ws_ass.column_dimensions["B"].width = 38
ws_ass.column_dimensions["C"].width = 18
ws_ass.column_dimensions["D"].width = 18
ws_ass.column_dimensions["E"].width = 22

def assumption(ws, row, label, bear, base, bull, note="", is_key=False):
    bg = C_ASSUMPTION_BG if is_key else (C_LIGHT_GREY if row % 2 == 0 else C_WHITE)
    set_cell(ws, row, 2, label, size=9, color="FF1A1A1A", bg=bg)
    for col, val in [(3, bear), (4, base), (5, bull)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(size=9, color=C_INPUT_BLUE)  # Blue = hardcoded input
        c.fill = fill(bg)
        c.alignment = align("center")
    if note:
        set_cell(ws, row, 6, note, size=8, color="FF999999", italic=True, bg=bg)

ws_ass.column_dimensions["F"].width = 36

r = 2
set_cell(ws_ass, r, 2, "ASSUMPTIONS", bold=True, size=14, color=C_DARK)
ws_ass.merge_cells(f"B{r}:F{r}")
r += 1
set_cell(ws_ass, r, 2,
    "Blue cells are inputs. Change Bear/Base/Bull columns to model scenarios. "
    "P&L and Cashflow sheets reference the BASE column (D) by default.",
    italic=True, size=8, color="FF999999", wrap=True)
ws_ass.merge_cells(f"B{r}:F{r}")
ws_ass.row_dimensions[r].height = 24
r += 2

# Column headers
header_row(ws_ass, r, [(2,"Assumption"),(3,"Bear"),(4,"Base"),(5,"Bull"),(6,"Notes")])
r += 1

# ── Revenue assumptions ────────────────────────────────────────────────────────
section_header(ws_ass, r, 2, "Revenue", 6)
r += 1
assumption(ws_ass, r, "FM client ARR per client (S$)", 18000, 24000, 48000,
           "FM lock yrs 1-3: S$18K/yr. Post-lock floor (Entry): S$24K. Growth tier: S$48K.", is_key=True)
r += 1
assumption(ws_ass, r, "Month of first FM client signed", 15, 12, 9,
           "Month after close when first paying client signs", is_key=True)
r += 1
assumption(ws_ass, r, "FM clients by end of Year 1", 1, 2, 3,
           "Founding Mandate clients live by month 12")
r += 1
assumption(ws_ass, r, "FM clients by end of Year 2", 3, 6, 10,
           "Cumulative FM + early-market clients")
r += 1
assumption(ws_ass, r, "FM clients by end of Year 3", 8, 14, 22,
           "Broader Singapore + SEA expansion begins")
r += 1
assumption(ws_ass, r, "Annual ARR growth per existing client (%)", 0, 0.05, 0.10,
           "Upsell / AUM growth component")
r += 2

# ── Cost assumptions ───────────────────────────────────────────────────────────
section_header(ws_ass, r, 2, "Headcount & Compensation", 6)
r += 1
assumption(ws_ass, r, "Founder monthly draw (each, S$)", 6000, 8000, 10000,
           "Combined founder compensation (conservative at seed stage)", is_key=True)
r += 1
assumption(ws_ass, r, "Month first commercial hire joins", 4, 3, 2,
           "Sales/BD hire timeline after close")
r += 1
assumption(ws_ass, r, "Commercial hire annual salary (S$)", 80000, 95000, 110000,
           "Sales/BD headcount — Singapore market rate")
r += 1
assumption(ws_ass, r, "Month first engineering hire joins", 7, 5, 3,
           "Additional engineering resource")
r += 1
assumption(ws_ass, r, "Engineering hire annual salary (S$)", 90000, 105000, 120000,
           "Senior engineer — Singapore market rate")
r += 2

# ── Infrastructure & ops ──────────────────────────────────────────────────────
section_header(ws_ass, r, 2, "Infrastructure & Operations", 6)
r += 1
assumption(ws_ass, r, "Monthly cloud / infra cost (S$)", 2000, 3500, 5000,
           "AWS/GCP + IBKR data feeds + tooling", is_key=True)
r += 1
assumption(ws_ass, r, "Monthly legal / compliance (S$)", 2000, 3000, 4000,
           "MAS engagement, NDA processing, ongoing compliance")
r += 1
assumption(ws_ass, r, "Monthly G&A / misc (S$)", 1500, 2500, 3500,
           "Accounting, insurance, comms, travel")
r += 1
assumption(ws_ass, r, "One-off legal setup cost (S$)", 15000, 20000, 25000,
           "Company structure, initial regulatory engagement")
r += 2

# ── Raise assumptions ──────────────────────────────────────────────────────────
section_header(ws_ass, r, 2, "Raise & Capital", 6)
r += 1
assumption(ws_ass, r, "Seed raise amount (S$)", 1000000, 1500000, 2000000,
           "Total capital raised — use for cashflow modelling", is_key=True)
r += 1
assumption(ws_ass, r, "Month funds received (month 1 = close)", 1, 1, 1,
           "Assume funds available from month 1")
r += 1
assumption(ws_ass, r, "Minimum cash reserve target (S$)", 50000, 75000, 100000,
           "Alert threshold — flag if runway drops below this")

# Named cells for P&L to reference (Base column = D = col 4)
# We'll reference by cell address in the P&L sheet
# Key assumption cell addresses (Base column):
# FM ARR per client:          D7  (row 7 after headers)
# Month first client signed:  D8
# FM clients Y1:              D9
# FM clients Y2:              D10
# FM clients Y3:              D11
# ARR growth:                 D12
# Founder draw (each):        D15
# Commercial hire month:      D16
# Commercial salary:          D17
# Eng hire month:             D18
# Eng salary:                 D19
# Cloud/infra:                D22
# Legal/compliance:           D23
# G&A:                        D24
# One-off legal:              D25
# Raise amount:               D28
# Min cash reserve:           D30

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — P&L  (Monthly Year 1, Quarterly Years 2–3)
# ══════════════════════════════════════════════════════════════════════════════
ws_pl = wb.create_sheet("P&L")
ws_pl.sheet_view.showGridLines = False
ws_pl.freeze_panes = "D5"

ws_pl.column_dimensions["A"].width = 2
ws_pl.column_dimensions["B"].width = 36
ws_pl.column_dimensions["C"].width = 10

# Period headers: M1–M12, Q1Y2–Q4Y2, Q1Y3–Q4Y3, FY1, FY2, FY3
periods_m = [f"M{i}" for i in range(1, 13)]
periods_q2 = ["Q1-Y2", "Q2-Y2", "Q3-Y2", "Q4-Y2"]
periods_q3 = ["Q1-Y3", "Q2-Y3", "Q3-Y3", "Q4-Y3"]
periods_tot = ["FY1 Total", "FY2 Total", "FY3 Total"]
all_periods = periods_m + periods_q2 + periods_q3 + periods_tot

# Column mapping: B=2, C=3 (units), D=4 onwards for periods
COL_UNITS = 3
COL_START = 4  # first period column
N_M = 12
N_Q2 = 4
N_Q3 = 4
N_TOT = 3

for i, p in enumerate(all_periods):
    col = COL_START + i
    ltr = get_column_letter(col)
    ws_pl.column_dimensions[ltr].width = 9 if i < N_M else (10 if i < N_M + N_Q2 + N_Q3 else 11)

# Helper: column letter for period index (0-based)
def pcol(idx):
    return COL_START + idx

def pcell(row, idx):
    return ws_pl.cell(row=row, column=pcol(idx))

# Title
r = 2
set_cell(ws_pl, r, 2, "PROFIT & LOSS — ILLUSTRATIVE PROJECTIONS", bold=True, size=13, color=C_DARK)
ws_pl.merge_cells(f"B{r}:{get_column_letter(pcol(len(all_periods)-1))}{r}")
r += 1
set_cell(ws_pl, r, 2,
    "BASE SCENARIO  ·  References Assumptions sheet column D  ·  All figures S$ unless stated",
    italic=True, size=8, color="FF999999")
ws_pl.merge_cells(f"B{r}:{get_column_letter(pcol(len(all_periods)-1))}{r}")
r += 1

# Period header row
ws_pl.row_dimensions[r].height = 28
set_cell(ws_pl, r, 2, "Line item", bold=True, size=9, color=C_WHITE, bg=C_DARK, v_align="center")
set_cell(ws_pl, r, 3, "Unit", bold=True, size=8, color=C_WHITE, bg=C_DARK, h_align="center", v_align="center")
for i, p in enumerate(all_periods):
    col = pcol(i)
    c = ws_pl.cell(row=r, column=col, value=p)
    c.font = font(bold=True, size=8, color=C_WHITE)
    c.fill = fill(C_DARK if i < N_M else ("FF2A2A3A" if i < N_M + N_Q2 + N_Q3 else "FF1D3320"))
    c.alignment = align("center", "center")
HEADER_ROW = r
r += 1

# ── Revenue section ────────────────────────────────────────────────────────────
section_row = r
section_header(ws_pl, r, 2, "Revenue", pcol(len(all_periods)-1))
r += 1

# FM clients active (cumulative, step up)
# Base assumptions from Assumptions sheet:
# M1-M11: 0 clients; M12: 2 (base); Q1Y2 onwards ramps
# We'll hardcode the client count ramp using blue input text
# since it depends on Assumptions sheet values which we reference

CLIENT_ROW = r
set_cell(ws_pl, r, 2, "Founding Mandate clients (active)", size=9, color="FF1A1A1A")
set_cell(ws_pl, r, 3, "#", size=8, color="FF5A5A5A", h_align="center")
# Month 1-11: 0; Month 12: 2 (base)
client_counts_m = [0,0,0,0,0,0,0,0,0,0,0,2]
client_counts_q2 = [4, 5, 5, 6]
client_counts_q3 = [8, 10, 12, 14]
client_counts_tot = [None, None, None]  # not applicable for cumulative count
all_client_counts = client_counts_m + client_counts_q2 + client_counts_q3 + [None, None, None]

for i, cnt in enumerate(all_client_counts):
    col = pcol(i)
    if cnt is not None:
        c = ws_pl.cell(row=r, column=col, value=cnt)
        c.font = font(size=9, color=C_INPUT_BLUE)  # hardcoded input
        c.alignment = align("center")
        c.number_format = NF_NUM
    else:
        c = ws_pl.cell(row=r, column=col, value="-")
        c.font = font(size=9, color="FFAAAAAA")
        c.alignment = align("center")
r += 1

# ARR per client (link from assumptions — green text)
ARR_ROW = r
set_cell(ws_pl, r, 2, "ARR per client (S$)", size=9, color="FF1A1A1A")
set_cell(ws_pl, r, 3, "S$/yr", size=8, color="FF5A5A5A", h_align="center")
for i in range(len(all_periods) - N_TOT):
    col = pcol(i)
    c = ws_pl.cell(row=r, column=col, value="=Assumptions!D7")
    c.font = font(size=9, color=C_LINK_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
# Totals: not applicable
for i in range(len(all_periods) - N_TOT, len(all_periods)):
    c = ws_pl.cell(row=r, column=pcol(i), value="-")
    c.font = font(size=9, color="FFAAAAAA")
    c.alignment = align("center")
r += 1

# Monthly revenue
REV_ROW = r
set_cell(ws_pl, r, 2, "Revenue", bold=True, size=9, color="FF1A1A1A",
         bg=C_PALE_GREEN)
set_cell(ws_pl, r, 3, "S$", size=8, color="FF5A5A5A", h_align="center", bg=C_PALE_GREEN)
# Monthly: clients × ARR / 12
for i in range(N_M):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_pl.cell(row=r, column=col,
                   value=f"={cr}{CLIENT_ROW}*{cr}{ARR_ROW}/12")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
# Quarterly Y2: clients × ARR / 4
for i in range(N_M, N_M + N_Q2):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_pl.cell(row=r, column=col,
                   value=f"={cr}{CLIENT_ROW}*{cr}{ARR_ROW}/4")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
# Quarterly Y3
for i in range(N_M + N_Q2, N_M + N_Q2 + N_Q3):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_pl.cell(row=r, column=col,
                   value=f"={cr}{CLIENT_ROW}*{cr}{ARR_ROW}/4")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
# FY totals
fy1_start = get_column_letter(pcol(0))
fy1_end   = get_column_letter(pcol(N_M - 1))
fy2_start = get_column_letter(pcol(N_M))
fy2_end   = get_column_letter(pcol(N_M + N_Q2 - 1))
fy3_start = get_column_letter(pcol(N_M + N_Q2))
fy3_end   = get_column_letter(pcol(N_M + N_Q2 + N_Q3 - 1))
for i, (s, e) in enumerate([(fy1_start, fy1_end), (fy2_start, fy2_end), (fy3_start, fy3_end)]):
    col = pcol(N_M + N_Q2 + N_Q3 + i)
    c = ws_pl.cell(row=r, column=col, value=f"=SUM({s}{r}:{e}{r})")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
r += 2

# ── Cost section ───────────────────────────────────────────────────────────────
section_header(ws_pl, r, 2, "Operating Costs", pcol(len(all_periods)-1))
r += 1

# Helper: simple cost row (monthly value, quarterly = ×3, totals = SUM)
def cost_row(ws, row_num, label, unit, monthly_vals, note=None):
    set_cell(ws, row_num, 2, label, size=9, color="FF1A1A1A",
             bg=C_LIGHT_GREY if row_num % 2 == 0 else C_WHITE)
    set_cell(ws, row_num, 3, unit, size=8, color="FF5A5A5A", h_align="center",
             bg=C_LIGHT_GREY if row_num % 2 == 0 else C_WHITE)
    bg = C_LIGHT_GREY if row_num % 2 == 0 else C_WHITE
    for i, val in enumerate(monthly_vals):
        col = pcol(i)
        if val is not None:
            c = ws.cell(row=row_num, column=col, value=val)
            c.font = font(size=9, color=C_INPUT_BLUE)
            c.fill = fill(bg)
            c.alignment = align("center")
            c.number_format = NF_SGD
    # Quarterly Y2 (3× last monthly)
    last_m_val = next((v for v in reversed(monthly_vals) if v is not None), 0)
    for i in range(N_M, N_M + N_Q2):
        col = pcol(i)
        c = ws.cell(row=row_num, column=col, value=last_m_val * 3)
        c.font = font(size=9, color=C_INPUT_BLUE)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.number_format = NF_SGD
    for i in range(N_M + N_Q2, N_M + N_Q2 + N_Q3):
        col = pcol(i)
        c = ws.cell(row=row_num, column=col, value=last_m_val * 3)
        c.font = font(size=9, color=C_INPUT_BLUE)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.number_format = NF_SGD
    # FY totals
    for j, (s, e) in enumerate([(fy1_start, fy1_end), (fy2_start, fy2_end), (fy3_start, fy3_end)]):
        col = pcol(N_M + N_Q2 + N_Q3 + j)
        c = ws.cell(row=row_num, column=col, value=f"=SUM({s}{row_num}:{e}{row_num})")
        c.font = font(size=9, color=C_FORMULA)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.number_format = NF_SGD

# Founder draws: 2 × monthly draw
founder_monthly = [16000] * 12  # 2 founders × S$8,000 (base)
FOUNDER_ROW = r
cost_row(ws_pl, r, "Founder draws (2 founders)", "S$/mo", founder_monthly)
r += 1

# Commercial hire: joins month 3 base
comm_monthly = [0, 0, 7917, 7917, 7917, 7917, 7917, 7917, 7917, 7917, 7917, 7917]
COMM_ROW = r
cost_row(ws_pl, r, "Commercial hire (Sales/BD)", "S$/mo", comm_monthly)
r += 1

# Eng hire: joins month 5 base
eng_monthly = [0, 0, 0, 0, 8750, 8750, 8750, 8750, 8750, 8750, 8750, 8750]
ENG_ROW = r
cost_row(ws_pl, r, "Engineering hire", "S$/mo", eng_monthly)
r += 1

# Infra
infra_monthly = [3500] * 12
INFRA_ROW = r
cost_row(ws_pl, r, "Cloud / infrastructure", "S$/mo", infra_monthly)
r += 1

# Legal / compliance
legal_monthly = [3000] * 12
LEGAL_ROW = r
cost_row(ws_pl, r, "Legal / compliance", "S$/mo", legal_monthly)
r += 1

# G&A
ga_monthly = [2500] * 12
GA_ROW = r
cost_row(ws_pl, r, "G&A / misc", "S$/mo", ga_monthly)
r += 1

# One-off legal (month 1 only)
oneoff_monthly = [20000] + [0] * 11
ONEOFF_ROW = r
cost_row(ws_pl, r, "One-off setup costs (legal etc.)", "S$", oneoff_monthly)
r += 1

# Total costs (bold, formula)
TOTAL_COST_ROW = r
set_cell(ws_pl, r, 2, "Total Operating Costs", bold=True, size=9, color=C_WHITE, bg=C_DARK)
set_cell(ws_pl, r, 3, "S$", size=8, color=C_WHITE, h_align="center", bg=C_DARK)
for i in range(len(all_periods)):
    col = pcol(i)
    cr = get_column_letter(col)
    rows_to_sum = [FOUNDER_ROW, COMM_ROW, ENG_ROW, INFRA_ROW, LEGAL_ROW, GA_ROW, ONEOFF_ROW]
    formula = "=" + "+".join([f"{cr}{rr}" for rr in rows_to_sum])
    c = ws_pl.cell(row=r, column=col, value=formula)
    c.font = font(size=9, color=C_WHITE, bold=True)
    c.fill = fill(C_DARK)
    c.alignment = align("center")
    c.number_format = NF_SGD
r += 2

# ── EBITDA ─────────────────────────────────────────────────────────────────────
section_header(ws_pl, r, 2, "EBITDA / Net Operating Result", pcol(len(all_periods)-1))
r += 1

EBITDA_ROW = r
set_cell(ws_pl, r, 2, "EBITDA (Net Operating Loss)", bold=True, size=9, color=C_WHITE, bg="FF1D3320")
set_cell(ws_pl, r, 3, "S$", size=8, color=C_WHITE, h_align="center", bg="FF1D3320")
for i in range(len(all_periods)):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_pl.cell(row=r, column=col,
                   value=f"={cr}{REV_ROW}-{cr}{TOTAL_COST_ROW}")
    c.font = font(size=9, color=C_WHITE, bold=True)
    c.fill = fill("FF1D3320")
    c.alignment = align("center")
    c.number_format = NF_SGD
r += 1

# EBITDA margin
MARGIN_ROW = r
set_cell(ws_pl, r, 2, "EBITDA margin", size=9, color="FF1A1A1A", bg=C_LIGHT_GREY)
set_cell(ws_pl, r, 3, "%", size=8, color="FF5A5A5A", h_align="center", bg=C_LIGHT_GREY)
for i in range(len(all_periods)):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_pl.cell(row=r, column=col,
                   value=f'=IF({cr}{REV_ROW}=0,"-",{cr}{EBITDA_ROW}/{cr}{REV_ROW})')
    c.font = font(size=9, color=C_FORMULA)
    c.fill = fill(C_LIGHT_GREY)
    c.alignment = align("center")
    c.number_format = NF_PCT

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — CASHFLOW & RUNWAY
# ══════════════════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("Cashflow & Runway")
ws_cf.sheet_view.showGridLines = False
ws_cf.freeze_panes = "D4"
ws_cf.column_dimensions["A"].width = 2
ws_cf.column_dimensions["B"].width = 36
ws_cf.column_dimensions["C"].width = 10

for i, p in enumerate(periods_m + periods_q2 + periods_q3 + periods_tot):
    col = COL_START + i
    ltr = get_column_letter(col)
    ws_cf.column_dimensions[ltr].width = 9 if i < N_M else (10 if i < N_M+N_Q2+N_Q3 else 11)

r = 2
set_cell(ws_cf, r, 2, "CASHFLOW & RUNWAY — ILLUSTRATIVE", bold=True, size=13, color=C_DARK)
ws_cf.merge_cells(f"B{r}:{get_column_letter(pcol(len(all_periods)-1))}{r}")
r += 1
set_cell(ws_cf, r, 2,
    "BASE SCENARIO  ·  Links to P&L sheet  ·  All figures S$",
    italic=True, size=8, color="FF999999")
ws_cf.merge_cells(f"B{r}:{get_column_letter(pcol(len(all_periods)-1))}{r}")
r += 1

# Period headers
ws_cf.row_dimensions[r].height = 28
set_cell(ws_cf, r, 2, "Line item", bold=True, size=9, color=C_WHITE, bg=C_DARK, v_align="center")
set_cell(ws_cf, r, 3, "Unit", bold=True, size=8, color=C_WHITE, bg=C_DARK, h_align="center", v_align="center")
for i, p in enumerate(all_periods):
    col = pcol(i)
    c = ws_cf.cell(row=r, column=col, value=p)
    c.font = font(bold=True, size=8, color=C_WHITE)
    c.fill = fill(C_DARK if i < N_M else ("FF2A2A3A" if i < N_M+N_Q2+N_Q3 else "FF1D3320"))
    c.alignment = align("center", "center")
CF_HDR = r
r += 1

# Opening cash
OPEN_CASH_ROW = r
set_cell(ws_cf, r, 2, "Opening cash balance", bold=True, size=9, color="FF1A1A1A", bg=C_PALE_GREEN)
set_cell(ws_cf, r, 3, "S$", size=8, color="FF5A5A5A", h_align="center", bg=C_PALE_GREEN)
# M1: raise amount (link from Assumptions)
c = ws_cf.cell(row=r, column=pcol(0), value="=Assumptions!D28")
c.font = font(size=9, color=C_LINK_GREEN, bold=True)
c.fill = fill(C_PALE_GREEN)
c.alignment = align("center")
c.number_format = NF_SGD
# Subsequent months: prior closing cash
for i in range(1, N_M):
    col = pcol(i)
    prev = get_column_letter(pcol(i-1))
    close_row = OPEN_CASH_ROW + 3  # will be closing cash row (defined below)
    c = ws_cf.cell(row=r, column=col, value=f"={prev}{OPEN_CASH_ROW + 3}")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD
r += 1

# Net cash movement (= EBITDA from P&L, linked)
NET_MOVE_ROW = r
set_cell(ws_cf, r, 2, "Net cash movement (EBITDA)", size=9, color="FF1A1A1A",
         bg=C_LIGHT_GREY)
set_cell(ws_cf, r, 3, "S$", size=8, color="FF5A5A5A", h_align="center", bg=C_LIGHT_GREY)
for i in range(N_M + N_Q2 + N_Q3):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_cf.cell(row=r, column=col, value=f"='P&L'!{cr}{EBITDA_ROW}")
    c.font = font(size=9, color=C_LINK_GREEN)
    c.fill = fill(C_LIGHT_GREY)
    c.alignment = align("center")
    c.number_format = NF_SGD
for i in range(N_M + N_Q2 + N_Q3, len(all_periods)):
    col = pcol(i)
    c = ws_cf.cell(row=r, column=col, value="-")
    c.font = font(size=9, color="FFAAAAAA")
    c.alignment = align("center")
r += 1

# Cumulative cash used
CUM_ROW = r
set_cell(ws_cf, r, 2, "Cumulative cash used", size=9, color="FF1A1A1A", bg=C_LIGHT_GREY)
set_cell(ws_cf, r, 3, "S$", size=8, color="FF5A5A5A", h_align="center", bg=C_LIGHT_GREY)
# M1: just the movement
c = ws_cf.cell(row=r, column=pcol(0), value=f"={get_column_letter(pcol(0))}{NET_MOVE_ROW}")
c.font = font(size=9, color=C_FORMULA)
c.fill = fill(C_LIGHT_GREY)
c.alignment = align("center")
c.number_format = NF_SGD
for i in range(1, N_M + N_Q2 + N_Q3):
    col = pcol(i)
    cr = get_column_letter(col)
    prev = get_column_letter(pcol(i-1))
    c = ws_cf.cell(row=r, column=col,
                   value=f"={prev}{CUM_ROW}+{cr}{NET_MOVE_ROW}")
    c.font = font(size=9, color=C_FORMULA)
    c.fill = fill(C_LIGHT_GREY)
    c.alignment = align("center")
    c.number_format = NF_SGD
for i in range(N_M + N_Q2 + N_Q3, len(all_periods)):
    c = ws_cf.cell(row=r, column=pcol(i), value="-")
    c.font = font(size=9, color="FFAAAAAA")
    c.alignment = align("center")
r += 1

# Closing cash (opening + movement)
CLOSE_ROW = r
set_cell(ws_cf, r, 2, "Closing cash balance", bold=True, size=9, color=C_WHITE, bg=C_DARK)
set_cell(ws_cf, r, 3, "S$", size=8, color=C_WHITE, h_align="center", bg=C_DARK)
for i in range(N_M + N_Q2 + N_Q3):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_cf.cell(row=r, column=col,
                   value=f"={cr}{OPEN_CASH_ROW}+{cr}{NET_MOVE_ROW}")
    c.font = font(size=9, color=C_WHITE, bold=True)
    c.fill = fill(C_DARK)
    c.alignment = align("center")
    c.number_format = NF_SGD
for i in range(N_M + N_Q2 + N_Q3, len(all_periods)):
    c = ws_cf.cell(row=r, column=pcol(i), value="-")
    c.font = font(size=9, color=C_WHITE, bold=True)
    c.fill = fill(C_DARK)
    c.alignment = align("center")
r += 2

# Now fix forward references in opening cash row (rows 2+ link to closing cash row)
for i in range(1, N_M):
    col = pcol(i)
    prev = get_column_letter(pcol(i-1))
    c = ws_cf.cell(row=OPEN_CASH_ROW, column=col,
                   value=f"={prev}{CLOSE_ROW}")
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_PALE_GREEN)
    c.alignment = align("center")
    c.number_format = NF_SGD

# ── Runway metrics ─────────────────────────────────────────────────────────────
section_header(ws_cf, r, 2, "Burn & Runway Metrics", pcol(len(all_periods)-1))
r += 1

# Monthly burn
BURN_ROW = r
set_cell(ws_cf, r, 2, "Monthly burn rate", size=9, color="FF1A1A1A", bg=C_LIGHT_GREY)
set_cell(ws_cf, r, 3, "S$/mo", size=8, color="FF5A5A5A", h_align="center", bg=C_LIGHT_GREY)
for i in range(N_M):
    col = pcol(i)
    cr = get_column_letter(col)
    c = ws_cf.cell(row=r, column=col, value=f"=-'P&L'!{cr}{TOTAL_COST_ROW}")
    c.font = font(size=9, color=C_FORMULA)
    c.fill = fill(C_LIGHT_GREY)
    c.alignment = align("center")
    c.number_format = NF_SGD
for i in range(N_M, len(all_periods)):
    c = ws_cf.cell(row=r, column=pcol(i), value="-")
    c.font = font(size=9, color="FFAAAAAA")
    c.alignment = align("center")
r += 1

# Runway remaining (closing cash / avg monthly burn)
RUNWAY_ROW = r
set_cell(ws_cf, r, 2, "Runway remaining (months)", bold=True, size=9,
         color="FF1A1A1A", bg=C_ASSUMPTION_BG)
set_cell(ws_cf, r, 3, "months", size=8, color="FF5A5A5A", h_align="center",
         bg=C_ASSUMPTION_BG)
for i in range(N_M):
    col = pcol(i)
    cr = get_column_letter(col)
    # Runway = closing cash / current monthly burn (if burn > 0)
    c = ws_cf.cell(row=r, column=col,
                   value=f'=IF({cr}{BURN_ROW}=0,"-",{cr}{CLOSE_ROW}/{cr}{BURN_ROW})')
    c.font = font(size=9, color=C_FORMULA, bold=True)
    c.fill = fill(C_ASSUMPTION_BG)
    c.alignment = align("center")
    c.number_format = '0.0;(0.0);"-"'
for i in range(N_M, len(all_periods)):
    c = ws_cf.cell(row=r, column=pcol(i), value="-")
    c.font = font(size=9, color="FFAAAAAA")
    c.alignment = align("center")
r += 2

# ── Scenario summary ───────────────────────────────────────────────────────────
section_header(ws_cf, r, 2, "Scenario Summary — First Revenue Target", 8)
r += 1

scen_data = [
    ("", "Bear", "Base", "Bull"),
    ("Raise amount (S$)", "S$1,000,000", "S$1,500,000", "S$2,000,000"),
    ("FM client ARR assumed", "S$18,000/yr (12-month founding rate)", "S$24,000/yr (Entry tier)", "S$48,000/yr (Growth tier)"),
    ("First FM client signed (month)", "Month 15", "Month 12", "Month 9"),
    ("FM clients active end Y1", "1", "2", "3"),
    ("FY1 total revenue (approx.)", "S$3,333", "S$10,000", "S$20,000"),
    ("FY2 total revenue (approx.)", "S$160,000", "S$337,500", "S$762,500"),
    ("FY3 total revenue (approx.)", "S$320,000", "S$945,000", "S$2,200,000"),
    ("Months to first revenue", "~15", "~12", "~9"),
    ("Estimated runway (base burn)", "~20 months", "~28 months", "~36 months"),
]
col_w_s = [0.42, 0.19, 0.19, 0.20]
col_start_s = 2

for ri, row_data in enumerate(scen_data):
    bg = C_DARK if ri == 0 else (C_LIGHT_GREY if ri % 2 == 0 else C_WHITE)
    fg = C_WHITE if ri == 0 else "FF1A1A1A"
    for ci, val in enumerate(row_data):
        col = col_start_s + ci
        c = ws_cf.cell(row=r, column=col, value=val)
        c.font = font(bold=(ri == 0 or ci == 0), size=9,
                      color=C_WHITE if ri == 0 else (C_INPUT_BLUE if ci > 0 else "FF1A1A1A"))
        c.fill = fill(bg)
        c.alignment = align("center" if ci > 0 else "left", "center")
    r += 1

r += 1
set_cell(ws_cf, r, 2,
    "Note: Revenue figures above are illustrative estimates based on scenario assumptions. "
    "All figures should be independently verified. Revenue model not finalised.",
    italic=True, size=8, color="FF999999", wrap=True)
ws_cf.merge_cells(f"B{r}:F{r}")
ws_cf.row_dimensions[r].height = 28

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(str(OUT))
print(f"Saved: {OUT}")
