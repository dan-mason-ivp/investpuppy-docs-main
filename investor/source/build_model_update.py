"""
InvestPuppy investor model update
- F2/F3: Update Assumptions (Tier A/B PP rates, raise params)
- F1:    Add Tenzor Scenario sheet
- All:   Update raise context across all scenario sheets + Cover
"""
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side, numbers)
from openpyxl.utils import get_column_letter

WB_IN  = '/home/claude/ip-investor-model.xlsx'
WB_OUT = '/home/claude/ip-investor-model.xlsx'

# ── Colour constants ──────────────────────────────────────────────────────────
BLUE_INPUT  = 'FF0000FF'   # blue  = hardcoded inputs
BLACK_CALC  = 'FF000000'   # black = formulas
GREEN_LINK  = 'FF008000'   # green = cross-sheet links
GREY_HDR    = 'FF404040'   # dark grey header text
DARK_BG     = 'FF0A0A0A'   # near-black background
GREEN_BG    = 'FF85D155'   # InvestPuppy green
LIGHT_BG    = 'FFF4F4F4'   # light section background
YELLOW_ATTN = 'FFFFFF00'   # yellow = needs attention

def ft(bold=False, color=BLACK_CALC, size=10, italic=False, name='Arial'):
    return Font(bold=bold, color=color, size=size,
                italic=italic, name=name)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def thin_border():
    s = Side(style='thin', color='FFD0D0D0')
    return Border(bottom=s)

def input_cell(ws, coord, value, fmt='#,##0;(#,##0);-'):
    ws[coord] = value
    ws[coord].font = ft(color=BLUE_INPUT)
    ws[coord].number_format = fmt

def calc_cell(ws, coord, formula):
    ws[coord] = formula
    ws[coord].font = ft(color=BLACK_CALC)

def link_cell(ws, coord, formula):
    ws[coord] = formula
    ws[coord].font = ft(color=GREEN_LINK)
    ws[coord].number_format = '#,##0;(#,##0);-'

def hdr(ws, coord, text, size=10, bold=True, bg=None, color=GREY_HDR):
    ws[coord] = text
    ws[coord].font = ft(bold=bold, color=color, size=size)
    if bg:
        ws[coord].fill = fill(bg)

def section(ws, coord, text):
    ws[coord] = text
    ws[coord].font = ft(bold=True, color=GREY_HDR, size=9)
    ws[coord].fill = fill('FFF0F0F0')

def note(ws, coord, text):
    ws[coord] = text
    ws[coord].font = ft(color='FF888888', size=8, italic=True)

wb = load_workbook(WB_IN)

# ═══════════════════════════════════════════════════════════════════════════════
# 1. COVER — update raise params + version + Tenzor note
# ═══════════════════════════════════════════════════════════════════════════════
cv = wb['Cover']
for row in cv.iter_rows():
    for cell in row:
        v = str(cell.value or '')
        if 'US$1.5M' in v:
            cell.value = v.replace('US$1.5M – US$2.25M', 'US$2M – US$2.5M')
        if 'US$6M' in v and 'US$10M' in v:
            cell.value = v.replace('US$6M – US$10M', 'US$9M – US$10M')
        if '1.1 · May 2026' in v:
            cell.value = v.replace('1.1 · May 2026', '1.2 · May 2026')

# Add Tenzor note to Cover
# Find last row with content
last_row = max(c.row for row in cv.iter_rows() for c in row if c.value)
r = last_row + 2
cv.cell(r, 2, 'Tenzor Scenario').font = ft(bold=True, color=GREY_HDR)
cv.cell(r, 3,
    'NDA-stage institutional product roadmap — combined Vektor + Tenzor scenario. '
    'Development begins post first Vektor Proof Partners go-live. Series A funded.'
).font = ft(color='FF888888', size=9)

# ═══════════════════════════════════════════════════════════════════════════════
# 2. ASSUMPTIONS — add Tier A/B PP rates, raise params
# ═══════════════════════════════════════════════════════════════════════════════
asm = wb['Assumptions']

# Update raise context references (rows 28-32 in scenario sheets reference strings)
# Update PP section header and add Tier A/B rates
# Current row 4: Proof Partners rate = 13500
# Add note and Tier B row
asm['B4'] = '  Proof Partners Tier A rate (<US$55M AUM)'
asm['B4'].font = ft(color=GREY_HDR, size=9)

# Insert a Tier B row — we'll add after row 4 by editing existing rows
# Actually we add new content at end of pricing section
# Row 9 is Bloomberg ref, row 10 is empty, row 11 is AUM thresholds
# Add Tier B between row 4 and 5
asm.insert_rows(5)
asm['B5'] = '  Proof Partners Tier B rate (>US$55M AUM, 35% below Growth tier)'
asm['B5'].font = ft(color=GREY_HDR, size=9)
for col, val in [('C', 23400), ('D', 23400), ('E', 23400)]:
    input_cell(asm, f'{col}5', val)
note(asm, 'F5', '=Growth tier × 0.65 · Year 1 preferential rate locked for duration')

# Add note on PP tier split below client ramp section
# Find the PP ramp rows (now shifted by 1 due to insert)
note(asm, 'F19', 'Split between Tier A and Tier B clients to be specified as PP conversations progress')

# Update raise params in scenario sheets
for sheet_name in ['Base Case', 'Bear Case', 'Bull Case']:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or '')
            if 'US$1.5M' in v:
                cell.value = v.replace('US$1.5M – US$2.25M', 'US$2M – US$2.5M')
            if 'US$6M – US$10M' in v:
                cell.value = v.replace('US$6M – US$10M', 'US$9M – US$10M')
            if '5-person team · institutional sales cycle · legal infrastructure · product build' in v:
                cell.value = ('5-person team · institutional sales cycle · '
                              'legal infrastructure · Tenzor design phase')

# ═══════════════════════════════════════════════════════════════════════════════
# 3. TENZOR SCENARIO — new sheet
# ═══════════════════════════════════════════════════════════════════════════════
if 'Tenzor Scenario' in wb.sheetnames:
    del wb['Tenzor Scenario']

tz = wb.create_sheet('Tenzor Scenario')

# Column widths
tz.column_dimensions['A'].width = 2
tz.column_dimensions['B'].width = 52
for col in ['C','D','E','F','G']:
    tz.column_dimensions[col].width = 16
tz.column_dimensions['H'].width = 32

YRS = ['Year 1\n(2026)', 'Year 2\n(2027)', 'Year 3\n(2028)',
       'Year 4\n(2029)', 'Year 5\n(2030)']
YCOLS = ['C','D','E','F','G']

# ── NDA header banner ─────────────────────────────────────────────────────────
tz['B1'] = 'TENZOR — INSTITUTIONAL PRODUCT ROADMAP'
tz['B1'].font = Font(bold=True, size=13, color='FFFFFFFF', name='Arial')
tz['B1'].fill = fill('FF0A0A0A')
tz['B1'].alignment = Alignment(vertical='center')
tz.row_dimensions[1].height = 24

tz['C1'] = 'NDA CONFIDENTIAL — NOT FOR EXTERNAL DISTRIBUTION'
tz['C1'].font = Font(bold=True, size=9, color='FF85D155', name='Arial')
tz['C1'].fill = fill('FF0A0A0A')
tz['C1'].alignment = Alignment(horizontal='center', vertical='center')

# ── Context ───────────────────────────────────────────────────────────────────
r = 3
tz[f'B{r}'] = 'TENZOR CONTEXT'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz[f'B{r}'].fill = fill('FFF0F0F0')

r += 1
ctx = [
    ('Product',         'Tenzor — full institutional portfolio management platform. '
                         'Same systematic methodology as Vektor extended to serve DPM desks '
                         'at regional banks, large MFOs, and institutional asset managers.'),
    ('Sequencing',      'Development begins post first Vektor Proof Partners go-live. '
                         'Series A funded. Not in development concurrently with Vektor.'),
    ('Series A trigger','First Vektor Proof Partners live · Alloy Partners Stage 2 active '
                         '· Tenzor design phase complete'),
    ('Series A size',   'US$4M – US$5M (indicative) · to fund Tenzor development team '
                         '+ expanded commercial infrastructure'),
    ('Market context',  'Temenos TripleA Plus: US$150K–US$400K+ per deployment. '
                         'Avaloq: US$500K–US$2M+ implementation. '
                         'Tenzor enters below both at mid-market institutional pricing.'),
    ('TAM (combined)',  '~US$25M–US$30M ARR at 10% penetration across SG, CH, UK '
                         '(Vektor + Tenzor non-overlapping, conservative)'),
]
for key, val in ctx:
    tz[f'B{r}'] = f'  {key}'
    tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
    tz[f'C{r}'] = val
    tz[f'C{r}'].font = ft(color='FF444444', size=9)
    r += 1

# ── Tenzor Assumptions ────────────────────────────────────────────────────────
r += 1
tz[f'B{r}'] = 'TENZOR ASSUMPTIONS — CONSERVATIVE BASE CASE'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz[f'B{r}'].fill = fill('FFF0F0F0')
r += 1

# Headers
for col, yr in zip(YCOLS, YRS):
    tz[f'{col}{r}'] = yr
    tz[f'{col}{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
    tz[f'{col}{r}'].alignment = Alignment(wrap_text=True, horizontal='center')
tz.row_dimensions[r].height = 28
r += 1

# Pricing assumptions
tz[f'B{r}'] = 'TENZOR PRICING (USD/year) — indicative'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=8)
r += 1

pricing = [
    ('  Tenzor Standard (up to US$250M AUM)', 120000),
    ('  Tenzor Growth (US$250M–US$1B AUM)',   180000),
    ('  Tenzor Institutional (US$1B+ AUM)',   280000),
]
for label, price in pricing:
    tz[f'B{r}'] = label
    tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
    for col in YCOLS:
        input_cell(tz, f'{col}{r}', price)
    note(tz, f'H{r}', 'Below Temenos/Avaloq floor · mid-market institutional pricing')
    r += 1

r += 1
# Client ramp
tz[f'B{r}'] = 'TENZOR CLIENT RAMP (cumulative, end of year)'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=8)
r += 1

note(tz, f'B{r}', 'Note: Development Years 1–3 (Vektor first). '
                   'Tenzor clients from Year 4 only. Numbers to be updated as Vektor Proof Partners '
                   'go-live and Series A timeline becomes clear.')
tz.row_dimensions[r].height = 14
r += 1

ramp_rows = [
    # label, Y1, Y2, Y3, Y4, Y5 (conservative — single-digit first year clients)
    ('  Tenzor Standard clients',    0, 0, 0, 1, 2),
    ('  Tenzor Growth clients',      0, 0, 0, 1, 2),
    ('  Tenzor Institutional clients',0, 0, 0, 0, 1),
]
ramp_start = r
for label, *vals in ramp_rows:
    tz[f'B{r}'] = label
    tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
    for col, val in zip(YCOLS, vals):
        input_cell(tz, f'{col}{r}', val)
    r += 1

# Additional costs from Series A (Year 3 onwards)
r += 1
tz[f'B{r}'] = 'TENZOR ADDITIONAL COSTS FROM SERIES A (USD/year)'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=8)
r += 1
cost_rows = [
    ('  Additional team (est. +8 people from Y3)',  0, 0, 840000, 840000, 960000),
    ('  Additional infrastructure',                  0, 0, 72000,  96000,  120000),
    ('  Legal + compliance (Tenzor-specific)',        0, 0, 80000,  60000,  60000),
    ('  Additional sales + marketing',               0, 0, 80000,  100000, 140000),
]
cost_start = r
for label, *vals in cost_rows:
    tz[f'B{r}'] = label
    tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
    for col, val in zip(YCOLS, vals):
        input_cell(tz, f'{col}{r}', val)
    r += 1
cost_end = r - 1

# ── Tenzor Revenue Model ──────────────────────────────────────────────────────
r += 1
tz[f'B{r}'] = 'TENZOR REVENUE (USD)'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz[f'B{r}'].fill = fill('FFF0F0F0')
r += 1

# We need to reference pricing and client count rows
# Pricing rows (Standard, Growth, Institutional): ramp_start-3, ramp_start-2, ramp_start-1
# Client ramp rows: ramp_start, ramp_start+1, ramp_start+2
price_rows = [ramp_start - 4, ramp_start - 3, ramp_start - 2]  # Standard, Growth, Institutional pricing
client_rows = [ramp_start, ramp_start + 1, ramp_start + 2]     # Standard, Growth, Institutional clients

tz_rev_labels = ['  Tenzor Standard revenue', '  Tenzor Growth revenue',
                  '  Tenzor Institutional revenue']
tz_rev_start = r
for i, (lbl, pr, cl) in enumerate(zip(tz_rev_labels, price_rows, client_rows)):
    tz[f'B{r}'] = lbl
    tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
    for col in YCOLS:
        calc_cell(tz, f'{col}{r}', f'={col}{pr}*{col}{cl}')
        tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    r += 1

tz_rev_end = r - 1
tz[f'B{r}'] = 'Tenzor Revenue Total'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz_total_rev_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}', f'=SUM({col}{tz_rev_start}:{col}{tz_rev_end})')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 1

# Tenzor additional costs total
r += 1
tz[f'B{r}'] = 'Tenzor Additional Costs Total'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz_cost_total_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}', f'=SUM({col}{cost_start}:{col}{cost_end})')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 1

# ── Combined Model ────────────────────────────────────────────────────────────
r += 2
tz[f'B{r}'] = 'COMBINED MODEL — VEKTOR (BASE CASE) + TENZOR'
tz[f'B{r}'].font = ft(bold=True, color='FFFFFFFF', size=10)
tz[f'B{r}'].fill = fill('FF0A0A0A')
tz[f'C{r}'] = '(Conservative · Base Vektor + Conservative Tenzor)'
tz[f'C{r}'].font = ft(bold=False, color='FF85D155', size=9)
tz[f'C{r}'].fill = fill('FF0A0A0A')

r += 2
# Column headers
for col, yr in zip(YCOLS, YRS):
    tz[f'{col}{r}'] = yr
    tz[f'{col}{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
    tz[f'{col}{r}'].alignment = Alignment(wrap_text=True, horizontal='center')
tz.row_dimensions[r].height = 28
r += 1

comb_sections = [
    # (label, formula_template, bold, color)
    ('Vektor Revenue (Base Case)',
     "='Base Case'!{col}21",  True,  GREEN_LINK),
    ('Tenzor Revenue',
     f'={{col}}{tz_total_rev_row}',  False, BLACK_CALC),
]
comb_rev_rows = []
for label, fmt, bold, col_color in comb_sections:
    tz[f'B{r}'] = label
    tz[f'B{r}'].font = ft(bold=bold, color=GREY_HDR, size=9)
    comb_rev_rows.append(r)
    for col in YCOLS:
        formula = fmt.replace('{col}', col)
        tz[f'{col}{r}'] = formula
        tz[f'{col}{r}'].font = ft(color=col_color)
        tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    r += 1

tz[f'B{r}'] = 'Total Combined Revenue'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=10)
comb_rev_total_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}',
              f'=SUM({col}{comb_rev_rows[0]}:{col}{comb_rev_rows[-1]})')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 2

# Costs
tz[f'B{r}'] = 'Vektor Operating Costs (Base Case)'
tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
vektor_cost_row = r
for col in YCOLS:
    link_cell(tz, f'{col}{r}', f"='Base Case'!{col}24")
r += 1

tz[f'B{r}'] = 'Tenzor Additional Costs'
tz[f'B{r}'].font = ft(color=GREY_HDR, size=9)
tenzor_cost_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}', f'={col}{tz_cost_total_row}')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
r += 1

tz[f'B{r}'] = 'Total Combined Costs'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
comb_cost_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}', f'={col}{vektor_cost_row}+{col}{tenzor_cost_row}')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 2

tz[f'B{r}'] = 'Combined EBITDA'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=10)
comb_ebitda_row = r
for col in YCOLS:
    calc_cell(tz, f'{col}{r}',
              f'={col}{comb_rev_total_row}-{col}{comb_cost_row}')
    tz[f'{col}{r}'].number_format = '#,##0;(#,##0);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 1

tz[f'B{r}'] = 'Combined EBITDA margin'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=10)
for col in YCOLS:
    calc_cell(tz, f'{col}{r}',
              f'=IF({col}{comb_rev_total_row}=0,"n/m",'
              f'{col}{comb_ebitda_row}/{col}{comb_rev_total_row})')
    tz[f'{col}{r}'].number_format = '0.0%;(0.0%);-'
    tz[f'{col}{r}'].font = ft(bold=True, color=BLACK_CALC)
r += 2

# ── Series A Context ──────────────────────────────────────────────────────────
tz[f'B{r}'] = 'SERIES A — CONTEXT'
tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
tz[f'B{r}'].fill = fill('FFF0F0F0')
r += 1

series_a = [
    ('Series A target',    'US$4M – US$5M (indicative) · to fund Tenzor development + expanded team'),
    ('Trigger milestone',  'First Vektor Proof Partners go-live · Alloy Partners Stage 2 active · Tenzor design phase complete'),
    ('Primary use',        'Tenzor development team · extended infrastructure · expanded sales & marketing'),
    ('Seed runway to A',   '18–24 months from seed close to Series A milestone'),
]
for key, val in series_a:
    tz[f'B{r}'] = f'  {key}'
    tz[f'B{r}'].font = ft(bold=True, color=GREY_HDR, size=9)
    tz[f'C{r}'] = val
    tz[f'C{r}'].font = ft(color='FF444444', size=9)
    r += 1

r += 1
note(tz, f'B{r}',
     'All Tenzor figures are indicative and will be updated following Vektor commercial launch '
     'and first Proof Partners go-live. Methodology, signal generation, and audit trail '
     'are identical in Vektor and Tenzor deployments.')

# ── Update Summary sheet ──────────────────────────────────────────────────────
sm = wb['Summary']
for row in sm.iter_rows():
    for cell in row:
        v = str(cell.value or '')
        if 'US$1.5M' in v:
            cell.value = v.replace('US$1.5M – US$2.25M', 'US$2M – US$2.5M')
        if 'US$6M – US$10M' in v:
            cell.value = v.replace('US$6M – US$10M', 'US$9M – US$10M')

# Add Tenzor line to Summary
last_sm = max(c.row for row in sm.iter_rows() for c in row if c.value)
r2 = last_sm + 2
sm[f'A{r2}'] = 'TENZOR SCENARIO (conservative base — see Tenzor Scenario tab)'
sm[f'A{r2}'].font = ft(bold=True, color=GREY_HDR, size=9)
r2 += 1
for label, formula_row in [
    ('  Combined Revenue (Vektor Base + Tenzor)', comb_rev_total_row),
    ('  Combined EBITDA',                          comb_ebitda_row),
]:
    sm[f'A{r2}'] = label
    sm[f'A{r2}'].font = ft(color=GREY_HDR, size=9)
    for col, yr_col in zip(['B','C','D','E','F'], YCOLS):
        link_cell(sm, f'{col}{r2}',
                  f"='Tenzor Scenario'!{yr_col}{formula_row}")
    r2 += 1

wb.save(WB_OUT)
print(f'Saved: {WB_OUT}')
print(f'Sheets: {wb.sheetnames}')
